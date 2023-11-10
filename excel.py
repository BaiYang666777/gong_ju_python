import openpyxl
from fuzzywuzzy import process

def open_and_modify_excel(file_path, sheet_name, cell_coordinates, new_values):
    try:
        # 打开 Excel 文件
        workbook = openpyxl.load_workbook(file_path)
        
        # 选择指定的工作表
        sheet = workbook[sheet_name]

        for i, cell_coordinate in enumerate(cell_coordinates):
            # 查询指定内容是否存在于 Excel 文件中
            if cell_coordinate not in sheet:
                print(f'单元格({cell_coordinate})不存在于工作表({sheet_name})中。')
                continue

            # 提取指定单元格的值
            cell_value = sheet[cell_coordinate].value
            print(f'原始值({cell_coordinate}): {cell_value}')

            # 修改单元格的值
            sheet[cell_coordinate] = new_values[i]

            print(f'已修改为({cell_coordinate}): {new_values[i]}')

        # 保存修改后的 Excel 文件
        workbook.save(file_path)
        
        print('保存成功')
        
    except Exception as e:
        print(f'发生错误: {e}')

def search_and_locate_value(file_path, sheet_name, search_value, fuzzy_search=False):
    try:
        # 打开 Excel 文件
        workbook = openpyxl.load_workbook(file_path)
        
        # 选择指定的工作表
        sheet = workbook[sheet_name]

        # 遍历工作表中所有单元格，查找指定内容
        if fuzzy_search:
            # 若进行模糊查询
            choices = [(cell.value, cell.coordinate) for row in sheet.iter_rows() for cell in row]
            matches = process.extract(search_value, choices, limit=1)
            if matches and matches[0][1] >= 80:  # 设置阈值，用于确定是否匹配
                match_value, match_coordinate = matches[0]
                print(f'找到模糊匹配内容({match_value})，位于单元格({match_coordinate}).')
                return match_coordinate
            else:
                print(f'未找到模糊匹配内容。')
                return None
        else:
            # 若进行精确查询
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value == search_value:
                        print(f'找到内容({search_value})，位于单元格({cell.coordinate})')
                        return cell.coordinate

            print(f'未找到内容({search_value})')
            return None
        
    except Exception as e:
        print(f'发生错误: {e}')
        return None

def print_menu():
    print("\n请选择操作:")
    print("1. 修改 Excel 文件")
    print("2. 查询内容并定位")
    print("3. 退出")

if __name__ == "__main__":
    print("欢迎使用 Excel 修改脚本！")
    
    while True:
        print_menu()
        
        choice = input("请输入操作编号: ")
        
        if choice == '1':
            # 获取用户输入的 Excel 文件路径
            file_path = input("请输入 Excel 文件路径：")

            # 获取用户输入的工作表名称
            sheet_name = input("请输入工作表名称：")

            # 获取用户输入的要修改的单元格坐标列表
            cell_coordinates = input("请输入要修改的单元格坐标列表（以逗号分隔）：").split(',')

            # 获取用户输入的新的值列表
            new_values = input("请输入新的值列表（以逗号分隔）：").split(',')

            # 调用函数进行操作
            open_and_modify_excel(file_path, sheet_name, cell_coordinates, new_values)
        
        elif choice == '2':
            # 获取用户输入的 Excel 文件路径
            file_path = input("请输入 Excel 文件路径：")

            # 获取用户输入的工作表名称
            sheet_name = input("请输入工作表名称：")

            # 获取用户输入的要查询的内容
            search_value = input("请输入要查询的内容：")

            # 获取用户是否进行模糊查询的选择
            fuzzy_search_choice = input("是否进行模糊查询？(y/n): ").lower()
            fuzzy_search = fuzzy_search_choice == 'y'

            # 调用函数进行查询和定位
            search_and_locate_value(file_path, sheet_name, search_value, fuzzy_search)

        elif choice == '3':
            print("谢谢使用，再见！")
            break
        
        else:
            print("无效的选择，请重新输入。")
